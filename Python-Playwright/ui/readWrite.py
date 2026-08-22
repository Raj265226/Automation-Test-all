import csv
import json
from openpyxl import load_workbook
import pytest
from playwright.sync_api import sync_playwright, expect


@pytest.fixture(scope='module')
def page():
    with sync_playwright() as p:
        browser = p.chromium.launch(
            executable_path=r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            headless=False,
            args=['--start-maximize']
        )
        context = browser.new_context()
        page = context.new_page()
        yield page
        browser.close()

def test_json_readwrite(page):
    with open('Files/ReadWrite/users.json', 'r') as file:
        users = json.load(file)
    for user in users:
        print('json->', user)
        page.goto('https://demoqa.com/text-box')
        page.locator('#userName').fill(user['username'])
        page.locator('#userEmail').fill(user['email'])
        page.locator('#submit').click()

        actual_username = page.locator('#output #name').text_content()
        actual_email = page.locator('#output #email').text_content()

        user['status'] = (
            'Pass'
            if actual_username.split(':')[1] == user['username']
            and actual_email.split(':')[1] == user['email']
            else 'Fail'
        )

    with open('Files/ReadWrite/users.json', 'w') as file:
        json.dump(users, file, indent=4)

def test_csv_readwrite(page):
    rows = []
    with open('Files/ReadWrite/users.csv', 'r') as file:
        users = csv.DictReader(file)
        for user in users:
            page.goto('https://demoqa.com/text-box')
            page.locator('#userName').fill(user['username'])
            page.locator('#userEmail').fill(user['email'])
            page.locator('#submit').click()

            actual_username = page.locator('#output #name').text_content()
            actual_email = page.locator('#output #email').text_content()

            user['status'] = (
                'Pass'
                if actual_username.split(':')[1] == user['username']
                and actual_email.split(':')[1] == user['email']
                else 'Fail'
            )
            rows.append(user)

    with open('Files/ReadWrite/users.csv', 'w', newline='') as file:
        updated_users = csv.DictWriter(file,fieldnames=['username', 'email', 'status'])
        updated_users.writeheader()
        updated_users.writerows(rows)

def test_excel_readwrite(page):
    workbook = load_workbook('Files/ReadWrite/users.xlsx')
    sheet = workbook.active
    headers = {}  # Create header dictionary

    for col in range(1, sheet.max_column + 1):
        # Read first row and build mapping
        header_name = sheet.cell(row=1, column=col).value
        headers[header_name] = col

    for row in range(2, sheet.max_row + 1):
        username = sheet.cell(row=row,column=headers['username']).value
        email = sheet.cell(row=row,column=headers['email']).value
        page.goto('https://demoqa.com/text-box')
        page.locator('#userName').fill(username)
        page.locator('#userEmail').fill(email)
        page.locator('#submit').click()

        actual_username = page.locator('#output #name').text_content()
        actual_email = page.locator('#output #email').text_content()

        status = (
            'Pass'
            if actual_username.split(':')[1] == username
            and actual_email.split(':')[1] == email
            else 'Fail'
        )

        sheet.cell(row=row,column=headers['status']).value = status
    workbook.save('Files/ReadWrite/users.xlsx')