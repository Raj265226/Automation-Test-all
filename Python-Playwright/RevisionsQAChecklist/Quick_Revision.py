from playwright.sync_api import sync_playwright, expect
import pytest
page, dialog, request, context, browser = None, None, None, None, None

# ============================================================

# Alerts
page.on("dialog", lambda dialog: dialog.accept())              # Normal alert
page.on("dialog", lambda dialog: dialog.dismiss())             # Cancel alert
page.on("dialog", lambda dialog: dialog.accept("Rohit"))       # Prompt alert
page.locator("#alertButton").click()                           # Click on alert
print("type:", dialog.type)                                    # Dialog type
print("message:", dialog.message)                              # Dialog message

# ============================================================

# API
response = request.get("/Library/GetBook.php",params={"AuthorName": "Rohit"}) # GET

payload = {
    "name": "Playwright",
    "isbn": "abcd",
    "aisle": "1234",
    "author": "Rohit"
}
response = request.post("/Library/Addbook.php",data=payload) # POST
response = request.put("https://reqres.in/api/users/2",data=payload) # PUT
response = request.delete("https://reqres.in/api/users/2") # DELETE
print(response.status)
print(response.json())

# ============================================================

# BROWSER ACTION==
page.goto("https://demoqa.com/alerts")                         # Open browser
page.reload()                                                  # Refresh
page.go_back()                                                 # Back
page.go_forward()                                              # Forward
print(page.url)
print(page.title())
page.close()                                                   # Close tab

# ============================================================

# BROWSER CONTEXT
context = browser.new_context(storage_state="xyz.json") # Context storage state

context1 = browser.new_context()                               # Context isolation
context2 = browser.new_context()

page1 = context.new_page()                                     # Page = tab
page2 = context.new_page()

# ============================================================

# CHECKBOX AND RADIO
page.get_by_role("checkbox",name="Select Desktop").click() # Single / Multiple checkbox
page.locator("#yesRadio").click()                              # Radio button

# ============================================================

# CROSS BROWSER
if browser_name == "chromium":
    browser = p.chromium.launch(
        executable_path="chrome.exe",
        headless=False,
        args=["--start-maximized"]
    )
elif browser_name == "firefox":
    browser = p.firefox.launch(headless=False)
else:
    raise ValueError(f"Unsupported browser: {browser_name}")

@pytest.mark.parametrize("browser_name",["chromium", "firefox"])
def test_browser_parameterize(my_browser):
    with sync_playwright() as p:
        browser = get_browser(p, my_browser)
        context = browser.new_context()
        page = context.new_page()
        context.close()
        browser.close()

# ============================================================

# DROPDOWN
page.locator("#oldSelectMenu").select_option("1")              # Value
page.locator("#oldSelectMenu").select_option(value="2")        # Value - another syntax
page.locator("#oldSelectMenu").select_option(label="Yellow")   # Label
page.locator("#oldSelectMenu").select_option(index=10)         # Index
page.locator("#cars").select_option(["volvo", "audi"])         # Multi-select

page.locator("#withOptGroup").click()                          # Open dropdown
page.get_by_text("Group 2, option 2").click()                  # Select option

page.locator("#react-select-2-input").fill("Another")          # Search
page.keyboard.press("Enter")                                   # Enter to select

# ============================================================

# DYNAMIC

page.locator("#visibleAfter").wait_for(state="visible")        # Wait until visible
page.locator("#xyz").wait_for(state="hidden")                  # Wait until hidden
page.locator("#xyz").wait_for(state="attached")                # Wait until attached
page.locator("#xyz").wait_for(state="detached")                # Wait until detached

enable_btn = page.locator("#enableAfter")
expect(enable_btn).to_be_enabled()
print("Enable button text:",enable_btn.inner_text())

paragraphs = page.locator(".js-scroll-added")
print(paragraphs.count())                                      # Calculate total


# ============================================================

# ELEMENT ACTION

page.locator("#doubleClickBtn").dblclick()                     # Double click
page.locator("#rightClickBtn").click(button="right")           # Right click
page.get_by_text("Click Me").last.click()                      # Dynamic ID / last is used

page.locator("#userName").fill("Rohit")                        # Enter data
page.locator("#userName").clear()                              # Clear data
page.locator("#userName").press("Control+A")                   # Keyboard combination
page.locator("#userName").press_sequentially("Rohit")          # Enter sequentially

page.get_by_text("Main Item 2").hover()                        # Mouse hover

source = page.locator("#draggable")                            # drag and drop
target = page.locator("#droppable").first
source.drag_to(target)

checkboxes = page.locator("input[type='checkbox']")            # check and uncheck
checkboxes.nth(0).check()
checkboxes.nth(1).uncheck()


# ============================================================

# ELEMENT VERIFY
fullname = page.locator("#userName")
print("Input value - ", fullname.input_value())                 # Check input value
print("Placeholder - ", fullname.get_attribute("placeholder")) # Get Attribute
print("Visible - ", fullname.is_visible())                     # Visible checking
print("Enable - ", fullname.is_enabled())                      # Enable checking
print("Disable - ", fullname.is_disabled())                    # Disable checking
print("Editable - ", fullname.is_editable())                   # Editable checking

message = page.locator("#dynamicClickMessage")                 # Retrieve Text Value
print("Text content - ", message.text_content())
print("Inner text - ", message.inner_text())
print("Inner html - ", message.inner_html())

yes_button = page.locator("#yesRadio")                         # Checkbox / Radio Verify
print("Checked - ", yes_button.is_checked())
print("Hidden - ", yes_button.is_hidden())

# ============================================================

# Exception Handling
try:
    page.locator("#wrongId").click()
except Exception as e:

    if type(e).__name__ == "TargetClosedError":
        print("Error A")                         # Page/context/browser closed

    elif type(e).__name__ == "TimeoutError":
        print("Error B")                         # Wait timeout

    elif type(e).__name__ == "Error":
        print("Strict mode / Playwright error")

    elif type(e).__name__ == "AssertionError":
        print("Assert Error")                    # Assertion failed

    else:
        print("Other Error:", e)

# ============================================================

# File Upload / Download
page.locator("#uploadFile").set_input_files("path/to/file.txt")  # File upload

with page.expect_download() as download_info:
    page.locator("#downloadButton").click()
download = download_info.value
download.save_as(f"Files_for_readwrite_n_upload_download/{download.suggested_filename}")
print("Saved - ", download.suggested_filename)
print("Download path - ", download.path())

assert download.path() is not None                       # Verify download

# ============================================================

# Frames
frame = page.frame_locator("#frame1")                    # Frame by ID
print(frame.locator("#sampleHeading").text_content())
frame = page.frame(name="iframeResult")                   # Frame by name
frame = page.frame(url="https://xyz.com")                 # Frame by URL

parent = page.frame(name="frame-top")                     # Parent / Child Frames
child1 = parent.child_frames[0]                           # Child using index
for frame in parent.child_frames:                         # Child using loop
    print("Frame name - ", frame.name)
    print("Frame URL - ", frame.url)
print("Child name - ", child1.name)
print("Parent URL - ", parent.url)

child = page.frame(name="frame-middle")                  # Switch from Child to Parent
print("Parent name - ", child.parent_frame.name)
print("Parent URL - ", child.parent_frame.url)

# ============================================================

# JavaScript Executor
title = page.evaluate("() => document.title")              # Title
handle = page.evaluate_handle("() => document.body")       # Handle
url = page.evaluate("() => window.location.href")          # URL
width = page.evaluate("() => window.innerWidth")           # Width

page.mouse.wheel(0, 100)                                        # Scroll little
page.evaluate("window.scrollTo(0, document.body.scrollHeight)") # Scroll bottom
page.evaluate("window.scrollTo(0, 0)")                          # Scroll top
page.locator("#submit").scroll_into_view_if_needed()            # Scroll until visible

textbox = page.locator("#userName")
textbox.evaluate("(element) => element.value = 'Rohit'")                 # JS - Enter Value
value = page.evaluate("() => document.querySelector('#userName').value") # Retrieve Value
print("Retrieve value - ", value)
button = page.locator("#submit")                                         # JS - Click
button.evaluate("(element) => element.click()")

# ============================================================

# Mouse / Keyboard Action
page.keyboard.press("R")                    # Press key
page.keyboard.type("oh")                    # Type
page.keyboard.down("Shift")                 # Press and hold
page.keyboard.up("Shift")                   # Release
page.keyboard.press("Control+A")            # Multiple key press

button = page.locator("#submit")
button.hover()                              # Mouse hover
page.mouse.down()                            # Press and hold mouse
page.mouse.up()                              # Release mouse

# ============================================================

# Parallel Execution
context1 = browser.new_context()        # Browser Isolation
context2 = browser.new_context()
page1 = context1.new_page()
page2 = context2.new_page()

# Run from terminal: pytest <path> -n 2 -v -s

context = browser.new_context()         # Context Isolation
page1 = context.new_page()
page2 = context.new_page()

# ============================================================

# Read / Write
import csv              # CSV read / write
rows = []
with open("users.csv", "r") as file:
    reader = csv.DictReader(file)
    for row in reader:
        page.goto("https://demoqa.com/text-box")
        username = row["username"]
        email = row["email"]
        page.locator("#userName").fill(username)
        page.locator("#userEmail").fill(email)
        actual_username = page.locator("#userName").input_value()
        actual_email = page.locator("#userEmail").input_value()
        row["status"] = (
            "Pass"
            if actual_username == username and actual_email == email
            else "Fail"
        )
        rows.append(row)
with open("users.csv", "w", newline="") as file:
    writer = csv.DictWriter(file,fieldnames=["username", "email", "status"])
    writer.writeheader()
    writer.writerows(rows)

import json                 # JSON read / write
with open("Files_for_readwrite_n_updownload/users.json","r") as file:
    reader = json.load(file)
    for usr in reader:
        page.goto("https://demoqa.com/text-box")
        page.locator("#userName").fill(usr["username"])
        page.locator("#userEmail").fill(usr["email"])
        actual_username = page.locator("#userName").input_value()
        actual_email = page.locator("#userEmail").input_value()
        usr["status"] = (
            "Pass"
            if actual_username == usr["username"]
            and actual_email == usr["email"]
            else "Fail"
        )
with open("Files_for_readwrite_n_updownload/users.json","w") as file:
    json.dump(reader, file, indent=4)

from openpyxl import load_workbook      # EXCEL read / write
workbook = load_workbook("Files/ReadWrite/users.xlsx")
sheet = workbook.active
headers = {}    # Read first row and build header mapping
for col in range(1, sheet.max_column + 1):
    header_name = sheet.cell(row=1,column=col).value
    headers[header_name] = col

for row in range(2, sheet.max_row + 1):
    username = sheet.cell(row=row,column=headers["username"]).value
    email = sheet.cell(row=row,column=headers["email"]).value
    page.goto("https://demoqa.com/text-box")
    page.locator("#userName").fill(username)
    page.locator("#userEmail").fill(email)
    page.locator("#submit").click()
    actual_username = page.locator("#output #name").text_content()
    actual_email = page.locator("#output #email").text_content()
    status = (
        "Pass"
        if actual_username.split(":")[1] == username
        and actual_email.split(":")[1] == email
        else "Fail"
    )
    sheet.cell(row=row,column=headers["status"]).value = status
workbook.save("Files/ReadWrite/users.xlsx")

# ============================================================

# Screenshot
page.screenshot(path="<file_path>")                         # Screenshot
page.screenshot(path="<file_path>", full_page=True)         # Full page Screenshot
username = page.locator("#userName")                        # Element Screenshot
username.screenshot(path="<file_path>")

try:                                                        # Screenshot on failure
    assert page.title() == "wrong title"
except:
    test_name = request.node.name
    page.screenshot(path=f"path/{test_name}.jpeg")

context = browser.new_context(record_video_dir="<path/>")    # Video download

context.tracing.start(                                       # Trace
    screenshots=True,
    snapshots=True,
    sources=True
)
page = context.new_page()
page.goto("https://demoqa.com/text-box")
context.tracing.stop( path="Files_for_readwrite_n_updownload/testing12.zip")

# ============================================================

# Shadow DOM
button = page.locator("#my-btn").first      # Shadow DOM / Playwright handles it

shadow_host = page.locator("#shadow-host")  # Inside Shadow Host
button = shadow_host.locator("#my-btn")

# ============================================================

# Wait
button = page.locator("#visibleAfter")      # Wait for visible
button.wait_for(state="visible")

page.wait_for_load_state("load")            # Wait for page load
page.wait_for_timeout(3000)                 # Wait for timeout
page.wait_for_load_state("domcontentloaded") # Wait for DOM content loaded
page.wait_for_function(                     # custom wait
    "() => document.querySelector('#enableAfter').disabled === false")
page.locator("#enableAfter").click()

# ============================================================

# Webtables
page.locator(".web-tables-wrapper table tbody tr").count()  # Row count
rows = page.locator(".web-tables-wrapper table tbody tr")   # Print all contents
for i in range(rows.count()):
    cols = rows.nth(i).locator("td").all_text_contents()
    cols.pop()
    print(" | ".join(cols))

for i in range(rows.count()):           # Read specific content
    row = rows.nth(i)
    if "Alden" in row.text_content():
        cols = row.locator("td").all_text_contents()
        cols.pop()
        print(" | ".join(cols))
        break

names = page.locator(                   # Sorting
".web-tables-wrapper table tbody tr td:nth-child(1)").all_text_contents()
assert names == sorted(names)

search_name = "Raj Roy"       # Search name until found
found = False
while True:
    rows = page.locator("#example tbody tr")
    for i in range(rows.count()):
        row = rows.nth(i)
        if search_name in row.locator("td").all_text_contents():
            found = True
            print("Found->"," | ".join(row.locator("td").all_text_contents()))
            break
    if found:
        break
    next_btn = page.locator("#example_next")
    if "disabled" in next_btn.get_attribute("class"):
        break
    next_btn.click()
    page.wait_for_timeout(1000)
try:
    assert found
except Exception as e:
    if type(e).__name__ == "AssertionError":
        print(f"{search_name} not found")

# ============================================================

# Windows and Tabs
page1 = context.new_page()                  # Multi Tabs
page2 = context.new_page()
page1.goto("https://demoqa.com/text-box")
page2.goto("https://demoqa.com/text-box")
print("total pages", len(context.pages))

with context.expect_page() as new_tab:      # New tab opens
    page1.locator("#tabButton").click()
child_tab = new_tab.value
print(child_tab.title(), child_tab.url)

with page1.expect_popup() as new_tab:       # Expect popup
    page1.locator("#tabButton").click()
child_tab = new_tab.value

with context.expect_page():                 # Switch by title
    page1.locator("#tabButton").click()
for p in context.pages:
    if p.title() == "DEMOQA":
        print("found", p.title())
        break

with context.expect_page() as info:         # Match by URL
    page1.locator("#tabButton").click()
tab = info.value
tab.wait_for_load_state()
for p in context.pages:
    if "sample" in p.url:
        print("matched", p.url)
        break

parent = context.new_page()                 # Close child
parent.goto("https://demoqa.com/browser-windows")
with context.expect_page() as child_info:
    parent.locator("#tabButton").click()
child = child_info.value
child.wait_for_load_state()
print("before close", len(context.pages))
child.close()
print("after close", len(context.pages))

# Parent -> child -> close all -> back to parent
parent = context.new_page()
parent.goto("https://demoqa.com/browser-windows")
with context.expect_page():
    parent.locator("#tabButton").click()
with context.expect_page():
    parent.locator("#windowButton").click()
with context.expect_page():
    parent.locator("#messageWindowButton").click()
print("before close tabs/windows",len(context.pages))

for p in context.pages:
    if p != parent:
        p.wait_for_load_state()
        print("url",p.url,"and title",p.title())
        p.close()

parent.bring_to_front()
print("after close tabs/windows",len(context.pages))