import pytest
import csv
import json
import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By


@pytest.fixture(scope="module")
def driver():
    driver = webdriver.Edge()
    driver.maximize_window()
    yield driver
    driver.quit()


def test_json_readwrite(driver):
    with open('Files/ReadWrite/users.json', 'r') as file:
        users = json.load(file)
    for user in users:
        print('json->', user)
        driver.get('https://demoqa.com/text-box')
        driver.find_element(By.ID, 'userName').send_keys(user['username'])
        driver.find_element(By.ID, 'userEmail').send_keys(user['email'])
        submit_btn = driver.find_element(By.ID, 'submit')
        driver.execute_script("arguments[0].click();", submit_btn)
        actual_username = driver.find_element(By.CSS_SELECTOR, '#output #name').text
        actual_email = driver.find_element(By.CSS_SELECTOR, '#output #email').text
        user['status'] = (
            'Pass'
            if actual_username.split(':')[1] == user['username']
            and actual_email.split(':')[1] == user['email']
            else 'Fail'
        )

    with open('Files/ReadWrite/users.json', 'w') as file:
        json.dump(users, file, indent=4)

def test_csv_readwrite(driver):
    rows = []
    with open('Files/ReadWrite/users.csv', 'r') as file:
        users = csv.DictReader(file)
        for user in users:
            driver.get('https://demoqa.com/text-box')
            driver.find_element(By.ID, 'userName').send_keys(user['username'])
            driver.find_element(By.ID, 'userEmail').send_keys(user['email'])
            submit_btn = driver.find_element(By.ID, 'submit')
            driver.execute_script("arguments[0].click();", submit_btn)
            actual_username = driver.find_element(By.CSS_SELECTOR, '#output #name').text
            actual_email = driver.find_element(By.CSS_SELECTOR, '#output #email').text
            user['status'] = (
                'Pass'
                if actual_username.split(':')[1] == user['username']
                and actual_email.split(':')[1] == user['email']
                else 'Fail'
            )
            rows.append(user)

    with open('Files/ReadWrite/users.csv','w',newline='') as file:
        updated_users = csv.DictWriter(file,fieldnames=['username', 'email', 'status'])
        updated_users.writeheader()
        updated_users.writerows(rows)

def test_excel_readwrite(driver):
    workbook = load_workbook('Files/ReadWrite/users.xlsx')
    sheet = workbook.active
    headers = {}  # Create header dictionary

    for col in range(1, sheet.max_column + 1):
        # Read first row and build mapping
        header_name = sheet.cell(row=1, column=col).value
        headers[header_name] = col

    for row in range(2, sheet.max_row + 1):
        username = sheet.cell(row=row,column=headers['username']).value
        email = sheet.cell(row=row,column=headers['email']).value
        driver.get('https://demoqa.com/text-box')
        driver.find_element(By.ID, 'userName').send_keys(username)
        driver.find_element(By.ID, 'userEmail').send_keys(email)
        submit_btn = driver.find_element(By.ID, 'submit')
        driver.execute_script("arguments[0].click();",submit_btn)
        actual_username = driver.find_element(By.CSS_SELECTOR,'#output #name').text
        actual_email = driver.find_element(By.CSS_SELECTOR,'#output #email').text
        status = (
            'Pass'
            if actual_username.split(':')[1] == username
            and actual_email.split(':')[1] == email
            else 'Fail'
        )
        sheet.cell(row=row,column=headers['status']).value = status
    workbook.save('Files/ReadWrite/users.xlsx')