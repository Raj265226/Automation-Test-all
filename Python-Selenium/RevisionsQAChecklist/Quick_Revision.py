from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import os
import pathlib
import time
import requests
import pytest

# ------------------------------------------------------------

# Alerts
alert = driver.switch_to.alert                         # normal alert
alert.accept()

alert = driver.switch_to.alert                         # cancel alert
alert.dismiss()

alert = driver.switch_to.alert                         # prompt alert
alert.send_keys('Rohit')
alert.accept()
print('Alert text ->', alert.text)                     # alert message

# ------------------------------------------------------------

# API
response = requests.get('/Library/GetBook.php', params={'AuthorName': 'Rohit'})   # Get

payload = {'name': 'Playwright', 'isbn': 'abcd', 'aisle': '1234', 'author': 'rohit'}
response = requests.post('/Library/Addbook.php', data=payload)                    # Post

response = requests.put('https://reqres.in/api/users/2', data=payload)             # Put

response = requests.delete('https://reqres.in/api/users/2')                        # Delete

print(response.status_code)
print(response.json())

# ------------------------------------------------------------

# Browser Action
driver.get("https://demoqa.com/alerts")                    # Open browser
driver.refresh()                                           # Refresh
driver.back()                                              # Back
driver.forward()                                           # Forward
driver.close()                                             # Close Tab
print(driver.current_url)
print(driver.title)

# ------------------------------------------------------------

# Browser Context
driver1, driver2 = drivers       # import drivers and in fixture there must be 2 drivers
driver1.get("https://demoqa.com/alerts")
driver2.get("https://demoqa.com/")

# Create new tab and again switch to main tab
driver.get("https://demoqa.com/alerts")
first_tab = driver.current_window_handle
driver.switch_to.new_window('tab')
driver.get("https://demoqa.com/")
driver.switch_to.window(first_tab)

# ------------------------------------------------------------

# Checkbox and Radio
driver.find_element(By.CSS_SELECTOR, "[aria-label='Select Desktop']").click()  # Single/Multi checkbox
driver.find_element(By.ID, 'yesRadio').click()                                 # Radio button

# ------------------------------------------------------------

# Cross Browser
def get_driver(browser_name):
    if browser_name == "chrome":
        driver = webdriver.Chrome()
    elif browser_name == "edge":
        driver = webdriver.Edge()
    else:
        raise ValueError(f"Unsupported browser -> {browser_name}")

    driver.maximize_window()
    return driver

@pytest.mark.parametrize("my_browser", ["chrome", "edge"])
def test_cross_browser(my_browser):
    driver = get_driver(my_browser)

    try:
        driver.get("https://demoqa.com/")
        print(f"\nBrowser -> {my_browser}")
        print("Title ->", driver.title)
        assert "DEMOSITE" in driver.title.upper()
    finally:
        driver.quit()

# ------------------------------------------------------------

# Dropdown
select_data = Select(driver.find_element(By.ID, 'oldSelectMenu'))   # select class
select_data.select_by_visible_text('Blue')                          # Visible text
select_data.select_by_index(3)                                     # Index
select_data.select_by_value('4')                                   # Value

select_data1 = Select(driver.find_element(By.ID, 'cars'))          # Multi select
select_data1.select_by_visible_text('Volvo')
select_data1.select_by_visible_text('Audi')

driver.find_element(By.ID, 'withOptGroup').click()                  # Open dropdown
driver.find_element(By.ID, 'react-select-2-input').send_keys(
    'Group 1, option 1', Keys.ENTER
)                                                                  # Select option

# ------------------------------------------------------------

# Dynamic
WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'XYZ'))) # Wait until enabled

WebDriverWait(driver, 10).until(EC.invisibility_of_element((By.ID, 'loading'))) # Wait until disappear

paragraphs = driver.find_elements(By.CLASS_NAME, 'jscroll-added')
print('Total stanzas -', len(paragraphs))                          # To calculate total

# ------------------------------------------------------------

# element action
double_click_btn = driver.find_element(By.ID, 'doubleClickBtn')    # Double click
ActionChains(driver).double_click(double_click_btn).perform()

right_click_btn = driver.find_element(By.ID, 'rightClickBtn')      # Right click
ActionChains(driver).context_click(right_click_btn).perform()

driver.find_element(By.ID, 'userName').send_keys('Rohit')           # Enter data
driver.find_element(By.ID, 'userName').clear()                     # Clear data
driver.find_element(By.ID, 'userName').send_keys(Keys.CONTROL, 'A') # Keyboard combination

main_item_2 = driver.find_element(By.LINK_TEXT, 'Main Item 2')
ActionChains(driver).move_to_element(main_item_2).perform()        # Mouse hover

source = driver.find_element(By.ID, 'draggable')                   # drag and drop
target = driver.find_elements(By.ID, 'droppable')[0]
ActionChains(driver).drag_and_drop(source, target).perform()

# ------------------------------------------------------------

# Element Verify

fullname = driver.find_element(By.ID, "userName")

print('Input value - ', fullname.get_attribute("value"))          # Check input value
print('Placeholder - ', fullname.get_attribute("placeholder"))   # Get Attribute
print('Visible - ', fullname.is_displayed())                     # Visible checking
print('Enable - ', fullname.is_enabled())                        # Enable checking

message = driver.find_element(By.ID, 'dynamicClickMessage')

print('Text content - ', message.text)                           # Retrieve Text value
print('Inner text - ',driver.execute_script("return arguments[0].textContent;", message))

print('Inner html - ',driver.execute_script("return arguments[0].innerHTML;", message))


# ------------------------------------------------------------------

# Exception

try:
    driver.get("https://demoqa.com/text-box")
    driver.find_element(By.ID, "wrongid")

except Exception as e:

    if type(e).__name__ == 'NoSuchElementException':
        print('Error A')                         # NoSuchElementException

    elif type(e).__name__ == 'TimeoutException':
        print('Error B')                         # wait for default time

    elif type(e).__name__ == 'StaleElementReferenceException':
        print('Error C')                         # DOM changed

    elif type(e).__name__ == 'AssertionError':
        print('Error D')                         # If assert fails

    elif type(e).__name__ == 'NoAlertPresentException':
        print('Error E')                         # If alert fails

    elif type(e).__name__ == 'NoSuchFrameException':
        print('Error F')                         # Frame fails to open

    elif type(e).__name__ == 'JavascriptException':
        print('Error G')                         # If JS fails

    elif type(e).__name__ == 'NoSuchWindowException':
        print('Error H')                         # Window fails


# ------------------------------------------------------------------

# File Upload / Download

file_path = os.path.abspath('Files/DownUpload/UploadFile.jpeg') # File upload

upload = driver.find_element(By.ID, "uploadFile")
upload.send_keys(file_path)

# Multiple file upload
driver.find_element(By.ID, "uploadFile").send_keys(f"{path1}\n{path2}")

# Download
driver.find_element(By.ID, 'downloadButton').click()


# ------------------------------------------------------------------

# Frames

frame = driver.find_element(By.ID, "frame1")       
driver.switch_to.frame(frame)                       # Frame by ID
print('Text value -->',driver.find_element(By.ID, "sampleHeading").text)

driver.get("https://www.w3schools.com/html/tryit.asp?filename=tryhtml_default") 
driver.switch_to.frame("iframeResult")              # Frame by name

# Nested / Child Frames
driver.switch_to.frame('frame-top')
child_frames = driver.find_elements(By.TAG_NAME, "frame")

for child_frame in child_frames:
    frame_name = child_frame.get_attribute("name")
    driver.switch_to.frame(child_frame)
    print("Inside ->", frame_name)
    driver.switch_to.parent_frame()                 # Switch to parent


# ------------------------------------------------------------------

# JS Executor

title = driver.execute_script("return document.title")       # Title
url = driver.execute_script("return window.location.href")    # URL
width = driver.execute_script("return window.innerWidth")     # Width

driver.execute_script("window.scrollBy(0,100)") # Scroll little
driver.execute_script("window.scrollTo(0,document.body.scrollHeight)") # Scroll bottom
driver.execute_script("window.scrollTo(0,0)")   # Scroll top
submit = driver.find_element(By.ID, "submit")   # Scroll until locator is found
driver.execute_script("arguments[0].scrollIntoView(true);",submit)

username = driver.find_element(By.ID, "userName")   # Enter value using JavaScript
driver.execute_script("arguments[0].value='Rohit';",username)

# Retrieve value
print('Retrieved value ->',driver.execute_script("return document.querySelector('#userName').value"))
driver.execute_script("arguments[0].click();",submit)   # Click using JavaScript

# ------------------------------------------------------------------

                        # ReadWrite file - CSV
import csv
rows = []
with open('Files/ReadWrite/users.csv', 'r') as file:
    users = csv.DictReader(file)
    for user in users:
        driver.get('https://demoqa.com/text-box')
        driver.find_element(By.ID, 'userName').send_keys(user['username'])
        driver.find_element(By.ID, 'userEmail').send_keys(user['email'])
        submit_btn = driver.find_element(By.ID, 'submit')
        driver.execute_script("arguments[0].click();", submit_btn)
        actual_username = driver.find_element(By.CSS_SELECTOR, '#output #name').text
        actual_email = driver.find_element(By.CSS_SELECTOR, '#output #email').text
        user['status'] = (
            'Pass'
            if actual_username.split(':')[1] == user['username']
            and actual_email.split(':')[1] == user['email']
            else 'Fail'
        )
        rows.append(user)

with open('Files/ReadWrite/users.csv','w',newline='') as file:
    updated_users = csv.DictWriter(file,fieldnames=['username', 'email', 'status'])
    updated_users.writeheader()
    updated_users.writerows(rows)

                            # json file
import json
with open('Files/ReadWrite/users.json', 'r') as file:
    users = json.load(file)
for user in users:
    print('json->', user)
    driver.get('https://demoqa.com/text-box')
    driver.find_element(By.ID, 'userName').send_keys(user['username'])
    driver.find_element(By.ID, 'userEmail').send_keys(user['email'])
    submit_btn = driver.find_element(By.ID, 'submit')
    driver.execute_script("arguments[0].click();", submit_btn)
    actual_username = driver.find_element(By.CSS_SELECTOR, '#output #name').text
    actual_email = driver.find_element(By.CSS_SELECTOR, '#output #email').text
    user['status'] = (
        'Pass'
        if actual_username.split(':')[1] == user['username']
        and actual_email.split(':')[1] == user['email']
        else 'Fail'
    )

with open('Files/ReadWrite/users.json', 'w') as file:
    json.dump(users, file, indent=4)

                        # Excel file
from openpyxl import load_workbook
workbook = load_workbook('Files/ReadWrite/users.xlsx')
sheet = workbook.active
headers = {}  # Create header dictionary

for col in range(1, sheet.max_column + 1):
    # Read first row and build mapping
    header_name = sheet.cell(row=1, column=col).value
    headers[header_name] = col

for row in range(2, sheet.max_row + 1):
    username = sheet.cell(row=row,column=headers['username']).value
    email = sheet.cell(row=row,column=headers['email']).value
    driver.get('https://demoqa.com/text-box')
    driver.find_element(By.ID, 'userName').send_keys(username)
    driver.find_element(By.ID, 'userEmail').send_keys(email)
    submit_btn = driver.find_element(By.ID, 'submit')
    driver.execute_script("arguments[0].click();",submit_btn)
    actual_username = driver.find_element(By.CSS_SELECTOR,'#output #name').text
    actual_email = driver.find_element(By.CSS_SELECTOR,'#output #email').text
    status = (
        'Pass'
        if actual_username.split(':')[1] == username
        and actual_email.split(':')[1] == email
        else 'Fail'
    )
    sheet.cell(row=row,column=headers['status']).value = status
workbook.save('Files/ReadWrite/users.xlsx')

# ------------------------------------------------------------------

# Screenshot

test_name = request.node.name
submit_btn = driver.find_element(By.ID, "submit")
submit_btn.screenshot(f'<path>')  # element Screenshot

submit_btn = driver.find_element(By.ID, "submit")  # Screenshot

submit_btn.screenshot('<file_path>')

try:                # screenshot at failure
    assert driver.title == 'wrong title'
except AssertionError:
    driver.save_screenshot("Files/DownUpload/failure.jpeg")

# ------------------------------------------------------------------

# ShadowDom

button = driver.find_elements(By.CSS_SELECTOR,'#my-btn')[0]  # shadow dom

shadow_host = driver.find_element(By.CSS_SELECTOR,'#shadow-host')  # Inside shadow host
shadow_root = shadow_host.shadow_root
btn2 = shadow_root.find_element(By.CSS_SELECTOR,'#my-btn')

# ------------------------------------------------------------------

# Wait
button = WebDriverWait(driver, 10).until(       # Wait for visible
    EC.visibility_of_element_located((By.ID, "visibleAfter")))


WebDriverWait(driver, 10).until(        # Wait for load
    lambda d: d.execute_script("return document.readyState") == "complete")

time.sleep(3)       # Wait for timeout

WebDriverWait(driver, 10).until(    # Custom wait
    lambda d: d.execute_script("return document.querySelector('#enableAfter').disabled === false"))
driver.find_element(By.ID,'enableAfter').click()

# ------------------------------------------------------------------

# WebTables
rows = driver.find_elements(By.CSS_SELECTOR,'.table tbody tr')  # row count
print('Total rows ->', len(rows))

for row in rows:        # read specific content
    if 'Alden' in row.text: # Find particular row
        col = row.find_elements(By.CSS_SELECTOR,'td')
        print(' | '.join(c.text for c in col[:-1]))
        break

# sorting
names = [   i.text for i in driver.find_elements(By.CSS_SELECTOR,'.table tbody tr td:nth-child(1)')]
try:
    assert names == sorted(names)

except AssertionError as e:
    print("Its not matching")
print('Names', names)
print('Sorted Names',sorted(names))

# Search name until found
search_name = 'Ethan Thomas'
found = False
while True:
    rows = driver.find_elements(By.CSS_SELECTOR,'#example tbody tr')
    for row in rows:
        if search_name in row.text:
            found = True
            print('Found->', ' | '.join(c.text for c in row.find_elements(By.TAG_NAME,'td')))
            break
    if found:
        break
    next_btn = driver.find_element(By.ID,'example_next')
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", next_btn)

    if 'disabled' in next_btn.get_attribute('class'):
        break

# ------------------------------------------------------------------

# Windows and Tabs
driver.get("https://demoqa.com")
driver.switch_to.new_window('tab')  # New tab opens
driver.get("https://demoqa.com/text-box")
print(len(driver.window_handles))

driver.get("https://demoqa.com/browser-windows")    # Multi tabs
tab1 = driver.current_window_handle
driver.find_element(By.CSS_SELECTOR,'#tabButton').click()
driver.switch_to.window(tab1)
print(len(driver.window_handles))

for window in driver.window_handles:    # switch by url
    driver.switch_to.window(window)
    if 'sample' in driver.current_url:
        break

for window in driver.window_handles:    # switch by title
    driver.switch_to.window(window)
    if driver.title == 'demosite':
        print('found ->',driver.title)
        break

# Parent to child and close all and back to parent
driver.get("https://demoqa.com/browser-windows")
parent_window = driver.current_window_handle
driver.find_element(By.ID,"tabButton").click()
driver.find_element(By.ID,"windowButton").click()
driver.find_element(By.ID,"messageWindowButton").click()
print('Before closing child tabs->',len(driver.window_handles))

for window in driver.window_handles:
    if window != parent_window:
        driver.switch_to.window(window)
        driver.close()
driver.switch_to.window(parent_window)
print('After closing child tabs->',len(driver.window_handles))










